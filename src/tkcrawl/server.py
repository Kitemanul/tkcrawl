"""知空 — FastAPI 后端服务"""

import asyncio
import json
import logging
from pathlib import Path

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from tkcrawl.auth import login_interactive
from tkcrawl.client import DouyinClient
from tkcrawl.models import VideoInfo
from tkcrawl.store import (
    save_comments,
    save_search_results,
    save_user_posts,
    save_user_profile,
    save_video,
)
from tkcrawl.utils import extract_aweme_id, extract_sec_user_id

logger = logging.getLogger("tkcrawl")

app = FastAPI(title="知空", version="0.1.0")

# ---- 状态管理 ----

_state = {
    "crawl_task": None,
    "crawl_status": "idle",  # idle | running | stopping
    "crawl_progress": [],
    "ws_clients": set(),
    "output_dir": "output",
}


async def broadcast(msg: dict):
    dead = set()
    for ws in _state["ws_clients"]:
        try:
            await ws.send_json(msg)
        except Exception:
            dead.add(ws)
    _state["ws_clients"] -= dead


async def log_progress(index: int, desc: str, extra: dict | None = None):
    entry = {"index": index, "desc": desc, **(extra or {})}
    _state["crawl_progress"].append(entry)
    await broadcast({"type": "video", **entry})


# ---- 数据接口 ----


@app.get("/api/videos")
async def list_videos(limit: int = 100, offset: int = 0):
    videos_dir = Path(_state["output_dir"]) / "videos"
    if not videos_dir.exists():
        return {"videos": [], "total": 0}

    files = sorted(
        videos_dir.glob("*.json"), key=lambda f: f.stat().st_mtime, reverse=True
    )
    total = len(files)
    files = files[offset : offset + limit]

    videos = []
    for f in files:
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            videos.append(data)
        except Exception:
            pass

    return {"videos": videos, "total": total}


@app.get("/api/videos/{aweme_id}")
async def get_video(aweme_id: str):
    path = Path(_state["output_dir"]) / "videos" / f"{aweme_id}.json"
    if not path.exists():
        return {"error": "not found"}, 404
    data = json.loads(path.read_text(encoding="utf-8"))
    return data


# ---- 登录与 Cookie ----


@app.get("/api/auth/status")
async def auth_status():
    cookie_path = Path("cookies/douyin_cookies.json")
    if cookie_path.exists():
        try:
            cookies = json.loads(cookie_path.read_text(encoding="utf-8"))
            names = {c["name"] for c in cookies}
            logged_in = "sessionid" in names
            return {"logged_in": logged_in, "cookie_count": len(cookies)}
        except Exception:
            pass
    return {"logged_in": False, "cookie_count": 0}


class LoginRequest(BaseModel):
    cookie_path: str = ""


@app.post("/api/auth/login")
async def start_login(req: LoginRequest):
    """启动扫码登录（会弹出浏览器窗口）"""
    if _state["crawl_status"] == "running":
        return {"error": "有任务在运行中，请先停止"}

    _state["crawl_status"] = "running"
    _state["crawl_progress"] = []

    async def do_login():
        try:
            await broadcast({"type": "video", "index": 1, "desc": "正在打开浏览器，请扫码登录..."})
            path = req.cookie_path or None
            success = await login_interactive(path)
            if success:
                await broadcast({"type": "done", "total": 1})
                await broadcast({"type": "video", "index": 2, "desc": "登录成功！Cookie 已保存"})
            else:
                await broadcast({"type": "error", "message": "登录超时，请重试"})
        except Exception as e:
            await broadcast({"type": "error", "message": str(e)})
        finally:
            _state["crawl_status"] = "idle"
            _state["crawl_task"] = None

    _state["crawl_task"] = asyncio.create_task(do_login())
    return {"status": "login_started"}


# ---- 采集控制 ----


class CrawlRequest(BaseModel):
    type: str = "feed"  # feed | video | user | comments | search
    max_count: int = 20
    headless: bool = True
    enrich_author: bool = True
    cookie_path: str = ""
    output: str = ""
    # feed 无额外参数
    # video / comments 需要 url
    url: str = ""
    # comments 额外参数
    with_replies: bool = False
    # search 参数
    keyword: str = ""
    search_type: str = "video"  # video | user


@app.post("/api/crawl/start")
async def start_crawl(req: CrawlRequest):
    if _state["crawl_status"] == "running":
        return {"error": "采集任务正在运行中"}

    _state["crawl_status"] = "running"
    _state["crawl_progress"] = []

    async def run_crawl():
        client = None
        seen_authors = set()
        cookie = req.cookie_path or None
        output_dir = req.output or _state["output_dir"]
        try:
            client = DouyinClient(headless=req.headless, cookie_path=cookie)
            await client.start()

            async def on_video(video: VideoInfo, index: int):
                if not video.is_valid:
                    logger.debug(f"跳过无效数据: {video.aweme_id}")
                    return
                if req.enrich_author and video.author.sec_uid:
                    if video.author.sec_uid not in seen_authors:
                        seen_authors.add(video.author.sec_uid)
                        await client.enrich_author(video)
                save_video(output_dir, video)
                await log_progress(index, video.desc[:60], {
                    "aweme_id": video.aweme_id,
                    "author": video.author.nickname,
                    "digg_count": video.stats.digg_count,
                    "cover_url": video.cover_url,
                })

            if req.type == "feed":
                await client.crawl_feed(
                    max_count=req.max_count, on_video=on_video
                )

            elif req.type == "video":
                aweme_id = extract_aweme_id(req.url)
                await log_progress(0, f"正在采集视频: {aweme_id}")
                info = await client.get_video(aweme_id)
                if req.enrich_author and info.author.sec_uid:
                    await client.enrich_author(info)
                save_video(output_dir, info)
                await log_progress(1, info.desc[:60], {
                    "aweme_id": info.aweme_id,
                    "author": info.author.nickname,
                    "digg_count": info.stats.digg_count,
                })

            elif req.type == "user":
                sec_uid = extract_sec_user_id(req.url)
                await log_progress(0, f"正在采集用户: {sec_uid[:30]}...")
                profile = await client.get_user_profile(sec_uid)
                save_user_profile(output_dir, profile)
                await log_progress(1, f"用户: {profile.nickname}  粉丝: {profile.follower_count}")
                posts = await client.get_user_posts(sec_uid, max_count=req.max_count)
                save_user_posts(output_dir, sec_uid, posts)
                for i, p in enumerate(posts):
                    save_video(output_dir, p)
                    await log_progress(i + 2, p.desc[:60], {
                        "aweme_id": p.aweme_id,
                        "author": p.author.nickname,
                        "digg_count": p.stats.digg_count,
                    })

            elif req.type == "comments":
                aweme_id = extract_aweme_id(req.url)
                await log_progress(0, f"正在采集评论: {aweme_id}")
                comment_list = await client.get_comments(
                    aweme_id,
                    max_count=req.max_count,
                    with_replies=req.with_replies,
                )
                save_comments(output_dir, aweme_id, comment_list)
                for i, c in enumerate(comment_list):
                    await log_progress(i + 1, f"{c.author_nickname}: {c.text[:50]}", {
                        "author": c.author_nickname,
                        "digg_count": c.like_count,
                    })

            elif req.type == "search":
                if not req.keyword:
                    await broadcast({"type": "error", "message": "请输入搜索关键词"})
                    return
                await log_progress(0, f"正在搜索: {req.keyword}")
                results = await client.search(
                    keyword=req.keyword,
                    search_type=req.search_type,
                    max_count=req.max_count,
                )
                for i, item in enumerate(results):
                    if isinstance(item, VideoInfo):
                        if not item.is_valid:
                            logger.debug(f"跳过无效搜索结果: {item.aweme_id}")
                            continue
                        if req.enrich_author and item.author.sec_uid:
                            if item.author.sec_uid not in seen_authors:
                                seen_authors.add(item.author.sec_uid)
                                await client.enrich_author(item)
                        save_video(output_dir, item)
                        await log_progress(i + 1, item.desc[:60], {
                            "aweme_id": item.aweme_id,
                            "author": item.author.nickname,
                            "digg_count": item.stats.digg_count,
                        })
                save_search_results(output_dir, req.keyword, results)

            await broadcast({
                "type": "done",
                "total": len(_state["crawl_progress"]),
            })
        except asyncio.CancelledError:
            await broadcast({"type": "stopped"})
        except Exception as e:
            logger.error(f"采集出错: {e}")
            await broadcast({"type": "error", "message": str(e)})
        finally:
            if client:
                await client.close()
            _state["crawl_status"] = "idle"
            _state["crawl_task"] = None

    _state["crawl_task"] = asyncio.create_task(run_crawl())
    return {"status": "started", "type": req.type}


# 保留旧路径兼容
@app.post("/api/crawl/feed")
async def start_crawl_feed(req: CrawlRequest):
    return await start_crawl(req)


@app.post("/api/crawl/stop")
async def stop_crawl():
    if _state["crawl_task"] and _state["crawl_status"] == "running":
        _state["crawl_status"] = "stopping"
        _state["crawl_task"].cancel()
        return {"status": "stopping"}
    return {"status": _state["crawl_status"]}


@app.get("/api/crawl/status")
async def crawl_status():
    return {
        "status": _state["crawl_status"],
        "count": len(_state["crawl_progress"]),
        "progress": _state["crawl_progress"],
    }


# ---- 导出 Excel ----


class ExportRequest(BaseModel):
    videos: list[dict]


@app.post("/api/export")
async def export_excel(req: ExportRequest):
    """将前端传入的筛选后视频数据导出为 Excel"""
    import io
    from datetime import datetime, timezone

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "视频数据"

    headers = [
        "视频ID", "标题", "作者", "作者主页", "粉丝数", "点赞数", "评论数",
        "分享数", "收藏数", "播放数", "话题标签", "发布时间",
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    for row, v in enumerate(req.videos, 2):
        stats = v.get("stats") or {}
        author = v.get("author") or {}
        create_time = v.get("create_time")
        time_str = ""
        if create_time:
            time_str = datetime.fromtimestamp(
                create_time, tz=timezone.utc
            ).strftime("%Y-%m-%d %H:%M")

        sec_uid = author.get("sec_uid", "")
        author_url = f"https://www.douyin.com/user/{sec_uid}" if sec_uid else ""

        ws.cell(row=row, column=1, value=v.get("aweme_id", ""))
        ws.cell(row=row, column=2, value=v.get("desc", ""))
        ws.cell(row=row, column=3, value=author.get("nickname", ""))
        ws.cell(row=row, column=4, value=author_url)
        ws.cell(row=row, column=5, value=author.get("follower_count", 0))
        ws.cell(row=row, column=6, value=stats.get("digg_count", 0))
        ws.cell(row=row, column=7, value=stats.get("comment_count", 0))
        ws.cell(row=row, column=8, value=stats.get("share_count", 0))
        ws.cell(row=row, column=9, value=stats.get("collect_count", 0))
        ws.cell(row=row, column=10, value=stats.get("play_count", 0))
        ws.cell(row=row, column=11, value=" ".join(v.get("hashtags") or []))
        ws.cell(row=row, column=12, value=time_str)

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # 标题列自动换行
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"zhikong_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---- WebSocket ----


@app.websocket("/api/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    _state["ws_clients"].add(ws)
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        pass
    finally:
        _state["ws_clients"].discard(ws)


# ---- 静态文件 (放最后，作为 catch-all) ----

_web_dir = Path(__file__).parent / "web"
if _web_dir.exists():
    app.mount("/", StaticFiles(directory=_web_dir, html=True), name="static")
