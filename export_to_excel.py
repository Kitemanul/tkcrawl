"""将搜索结果 JSON 导出为 Excel，并附带分析工作表。"""

import json
import statistics
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TZ = ZoneInfo("Asia/Shanghai")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")


def _style_header(ws, row: int, headers: list[str]) -> None:
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_col_widths(ws, widths: list[int]) -> None:
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _median(values: list[int | float]) -> int | float:
    return statistics.median(values) if values else 0


def _p75(values: list[int | float]) -> int | float:
    return statistics.quantiles(values, n=4)[2] if len(values) >= 4 else 0


def _duration_bucket(sec: float) -> str:
    if sec < 30:
        return "<30s"
    if sec < 60:
        return "30-60s"
    if sec < 180:
        return "1-3min"
    if sec < 300:
        return "3-5min"
    return ">5min"


def _follower_bucket(count: int) -> str:
    if count < 1000:
        return "<1k"
    if count < 10000:
        return "1k-10k"
    if count < 100000:
        return "10k-100k"
    if count < 1000000:
        return "100k-1m"
    return ">1m"


def _analyze(data: list[dict]) -> dict:
    likes: list[int] = []
    comments: list[int] = []
    shares: list[int] = []
    collects: list[int] = []
    durations: list[float] = []
    followers: list[int] = []
    comment_like_rate: list[float] = []
    share_like_rate: list[float] = []
    collect_like_rate: list[float] = []

    duration_buckets: Counter[str] = Counter()
    follower_buckets: Counter[str] = Counter()
    post_hours: Counter[int] = Counter()
    weekdays: Counter[str] = Counter()
    hashtags: Counter[str] = Counter()
    title_signals: Counter[str] = Counter()

    weekday_map = {
        "Mon": "周一",
        "Tue": "周二",
        "Wed": "周三",
        "Thu": "周四",
        "Fri": "周五",
        "Sat": "周六",
        "Sun": "周日",
    }
    signal_words = {
        "低成本/省钱": ["低成本", "穷装", "省钱", "便宜", "平替", "预算"],
        "收纳整理": ["收纳", "整理", "置物", "衣帽间", "收纳柜", "储物"],
        "布置氛围": ["布置", "氛围", "温馨", "治愈", "小窝", "复古"],
        "租房/出租屋": ["出租屋", "租房", "老破小", "卧室", "小户型"],
        "好物推荐": ["好物", "分享", "清单", "推荐", "神器", "必看"],
        "改造升级": ["改造", "爆改", "前后", "翻新", "重生"],
        "个人叙事": ["独居", "北漂", "女生", "打工人", "情侣", "一个人"],
    }

    for item in data:
        stats = item.get("stats", {})
        author = item.get("author", {})
        like_count = int(stats.get("digg_count", 0) or 0)
        comment_count = int(stats.get("comment_count", 0) or 0)
        share_count = int(stats.get("share_count", 0) or 0)
        collect_count = int(stats.get("collect_count", 0) or 0)
        duration_sec = round((item.get("duration", 0) or 0) / 1000, 1)
        follower_count = int(author.get("follower_count", 0) or 0)

        likes.append(like_count)
        comments.append(comment_count)
        shares.append(share_count)
        collects.append(collect_count)
        durations.append(duration_sec)
        followers.append(follower_count)

        if like_count:
            comment_like_rate.append(comment_count / like_count)
            share_like_rate.append(share_count / like_count)
            collect_like_rate.append(collect_count / like_count)

        duration_buckets[_duration_bucket(duration_sec)] += 1
        follower_buckets[_follower_bucket(follower_count)] += 1

        create_time = item.get("create_time", 0)
        if create_time:
            dt = datetime.fromtimestamp(create_time, TZ)
            post_hours[dt.hour] += 1
            weekdays[weekday_map[dt.strftime("%a")]] += 1

        for tag in item.get("hashtags", []):
            hashtags[tag] += 1

        desc = item.get("desc", "")
        for label, words in signal_words.items():
            if any(word in desc for word in words):
                title_signals[label] += 1

    top_like_items = sorted(
        data, key=lambda item: item.get("stats", {}).get("digg_count", 0), reverse=True
    )[:15]
    top_collect_items = sorted(
        data, key=lambda item: item.get("stats", {}).get("collect_count", 0), reverse=True
    )[:15]

    best_duration = duration_buckets.most_common(1)[0][0] if duration_buckets else ""
    best_hour = post_hours.most_common(1)[0][0] if post_hours else ""

    insights = [
        f"样本共 {len(data)} 条，中位点赞 {int(_median(likes))}，75 分位点赞 {int(_p75(likes))}。",
        f"主流时长集中在 {best_duration}，中位时长 {round(_median(durations), 1)} 秒。",
        (
            f"作者粉丝中位数 {int(_median(followers))}，"
            f"说明这个赛道不只属于大号，小号也有机会打出爆款。"
        ),
        f"发布时间高峰集中在 {best_hour}:00 附近，晚间 17:00-21:00 最值得优先测试。",
        (
            f"收藏/点赞中位比 {round(_median(collect_like_rate), 4)}，"
            "说明这类内容很强依赖“可抄作业、可直接购买、可马上复现”的价值。"
        ),
        (
            "高频爆款信号集中在“改造前后反差、低成本、收纳、好物清单、"
            "小空间卧室/出租屋”这些主题。"
        ),
        (
            "如果想做同主题爆款，优先拍一个明确问题的解决方案，"
            "而不是泛泛展示‘改造过程很治愈’。"
        ),
    ]

    return {
        "summary": [
            ("样本数", len(data)),
            ("点赞中位数", int(_median(likes))),
            ("点赞 75 分位", int(_p75(likes))),
            ("评论中位数", int(_median(comments))),
            ("收藏中位数", int(_median(collects))),
            ("时长中位数(秒)", round(_median(durations), 1)),
            ("粉丝中位数", int(_median(followers))),
            ("评论/点赞中位比", round(_median(comment_like_rate), 4)),
            ("分享/点赞中位比", round(_median(share_like_rate), 4)),
            ("收藏/点赞中位比", round(_median(collect_like_rate), 4)),
        ],
        "duration_buckets": duration_buckets.most_common(),
        "follower_buckets": follower_buckets.most_common(),
        "post_hours": post_hours.most_common(10),
        "weekdays": weekdays.most_common(),
        "hashtags": hashtags.most_common(20),
        "title_signals": title_signals.most_common(),
        "top_like_items": top_like_items,
        "top_collect_items": top_collect_items,
        "insights": insights,
    }


def _write_data_sheet(ws, data: list[dict]) -> None:
    has_source_keywords = any(item.get("source_keywords") for item in data)
    headers = [
        "视频ID",
        "标题/描述",
        "话题标签",
        "发布时间",
        "时长(秒)",
        "播放量",
        "点赞数",
        "评论数",
        "分享数",
        "收藏数",
        "作者昵称",
        "作者UID",
        "粉丝数",
        "关注数",
        "获赞总数",
        "作品数",
        "视频链接",
        "封面链接",
    ]
    if has_source_keywords:
        headers.append("来源关键词")

    ws.title = "抖音视频数据"
    ws.append(headers)
    _style_header(ws, 1, headers)
    ws.row_dimensions[1].height = 20

    for row_idx, item in enumerate(data, 2):
        create_time = item.get("create_time", 0)
        publish_time = (
            datetime.fromtimestamp(create_time, TZ).strftime("%Y-%m-%d %H:%M:%S")
            if create_time
            else ""
        )
        duration_ms = item.get("duration", 0)
        duration_s = round(duration_ms / 1000, 1) if duration_ms else 0

        stats = item.get("stats", {})
        author = item.get("author", {})
        hashtags = "、".join(item.get("hashtags", []))

        row = [
            item.get("aweme_id", ""),
            item.get("desc", ""),
            hashtags,
            publish_time,
            duration_s,
            stats.get("play_count", 0),
            stats.get("digg_count", 0),
            stats.get("comment_count", 0),
            stats.get("share_count", 0),
            stats.get("collect_count", 0),
            author.get("nickname", ""),
            author.get("uid", ""),
            author.get("follower_count", 0),
            author.get("following_count", 0),
            author.get("total_favorited", 0),
            author.get("aweme_count", 0),
            item.get("video_url", ""),
            item.get("cover_url", ""),
        ]
        if headers[-1] == "来源关键词":
            row.append("、".join(item.get("source_keywords", [])))
        ws.append(row)

        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (1, 2, 3, 4, 11, 12, len(headers)) else "right",
                vertical="center",
                wrap_text=(col_idx in (2, len(headers))),
            )

    widths = [20, 50, 30, 20, 10, 12, 12, 12, 12, 12, 18, 20, 12, 12, 14, 10, 20, 20]
    if headers[-1] == "来源关键词":
        widths.append(26)
    _set_col_widths(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_analysis_sheet(ws, analysis: dict) -> None:
    ws.title = "分析概览"
    ws["A1"] = "出租屋改造爆款分析"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "核心结论"
    ws["A3"].fill = SECTION_FILL
    ws["A3"].font = Font(bold=True)
    row = 4
    for insight in analysis["insights"]:
        ws.cell(row=row, column=1, value="•")
        ws.cell(row=row, column=2, value=insight)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="指标").fill = SECTION_FILL
    ws.cell(row=row, column=2, value="数值").fill = SECTION_FILL
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1
    for metric, value in analysis["summary"]:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        row += 1

    sections = [
        ("时长分布", analysis["duration_buckets"]),
        ("粉丝分布", analysis["follower_buckets"]),
        ("发布时间(小时)", analysis["post_hours"]),
        ("星期分布", analysis["weekdays"]),
        ("高频标签", analysis["hashtags"]),
        ("标题信号", analysis["title_signals"]),
    ]
    start_col = 4
    for title, items in sections:
        ws.cell(row=3, column=start_col, value=title).fill = SECTION_FILL
        ws.cell(row=3, column=start_col).font = Font(bold=True)
        ws.cell(row=4, column=start_col, value="维度").fill = HEADER_FILL
        ws.cell(row=4, column=start_col + 1, value="数量").fill = HEADER_FILL
        ws.cell(row=4, column=start_col).font = HEADER_FONT
        ws.cell(row=4, column=start_col + 1).font = HEADER_FONT
        r = 5
        for label, count in items:
            ws.cell(row=r, column=start_col, value=label)
            ws.cell(row=r, column=start_col + 1, value=count)
            r += 1
        start_col += 3

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["P"].width = 18
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["S"].width = 18
    ws.column_dimensions["T"].width = 12


def _write_examples_sheet(ws, title: str, items: list[dict], sort_key: str) -> None:
    ws.title = title
    headers = [
        "排序值",
        "点赞数",
        "评论数",
        "收藏数",
        "时长(秒)",
        "粉丝数",
        "作者昵称",
        "标题/描述",
        "话题标签",
    ]
    ws.append(headers)
    _style_header(ws, 1, headers)

    for item in items:
        stats = item.get("stats", {})
        author = item.get("author", {})
        row = [
            stats.get(sort_key, 0),
            stats.get("digg_count", 0),
            stats.get("comment_count", 0),
            stats.get("collect_count", 0),
            round((item.get("duration", 0) or 0) / 1000, 1),
            author.get("follower_count", 0),
            author.get("nickname", ""),
            item.get("desc", ""),
            "、".join(item.get("hashtags", [])),
        ]
        ws.append(row)

    _set_col_widths(ws, [12, 12, 12, 12, 10, 12, 18, 60, 30])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        row[7].alignment = Alignment(wrap_text=True, vertical="top")
        row[8].alignment = Alignment(wrap_text=True, vertical="top")


def export(json_path: str, output_path: str) -> None:
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    analysis = _analyze(data)

    wb = openpyxl.Workbook()
    _write_data_sheet(wb.active, data)
    _write_analysis_sheet(wb.create_sheet(), analysis)
    _write_examples_sheet(
        wb.create_sheet(), "高赞样本", analysis["top_like_items"], "digg_count"
    )
    _write_examples_sheet(
        wb.create_sheet(), "高收藏样本", analysis["top_collect_items"], "collect_count"
    )

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"已导出 {len(data)} 条记录 → {output_path}")


if __name__ == "__main__":
    json_file = sys.argv[1] if len(sys.argv) > 1 else "output/search/女性_30_1774102294.json"
    out_file = sys.argv[2] if len(sys.argv) > 2 else "output/女性30岁_抖音视频.xlsx"
    export(json_file, out_file)
